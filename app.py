from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session, send_file, g
import sqlite3
from datetime import datetime, time
import requests
import base64
import dlib 
from scipy.spatial import distance as dist
import numpy as np
import cv2
import face_recognition
import pickle
import os
import time
from io import BytesIO
import pytz
import mediapipe as mp
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash
from flask_socketio import SocketIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

app = Flask(__name__)
app.secret_key = os.environ.get('FLASK_SECRET_KEY', 'a_default_super_secret_key_that_should_be_changed_2025')

# Flask Session Configuration
app.config['SESSION_COOKIE_PATH'] = '/'
app.config['SESSION_COOKIE_NAME'] = 'my_attendance_session'
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SECURE'] = False # Set to True in production with HTTPS
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'

# Initialize Flask-SocketIO
socketio = SocketIO(app, cors_allowed_origins="*", async_mode='threading')

DB_FILE = 'attendance.db'
ENCODINGS_PATH = 'known_face_encodings.pkl'

# Global lists/dictionaries for in-memory storage of known faces and cooldowns
known_face_encodings = []
known_face_roll_numbers = []
last_logged_time = {} # Cooldown for attendance logging per student
COOLDOWN_PERIOD = 30 # Seconds: How long before a student can be logged again
AGE_MOOD_COOLDOWN_PERIOD = 5 # Seconds: How often to run age/mood detection
last_age_mood_time = {} # Cooldown for age/mood detection per recognized face (even unknowns)
last_liveness_check_time = 0 
LIVENESS_CHECK_COOLDOWN = 10 

# Initialize MediaPipe Face Detection
face_detector = None
try:
    face_detector = mp.solutions.face_detection.FaceDetection(model_selection=0, min_detection_confidence=0.5)
    print("MediaPipe FaceDetection initialized.")
except Exception as e:
    print(f"Error initializing MediaPipe FaceDetection: {e}")
    print("Please ensure MediaPipe is correctly installed and its face_detection solution is available.")
    face_detector = None

MODEL_DIR = 'models'
AGE_MODEL_PROTO = os.path.join(MODEL_DIR, 'deploy_age.prototxt')
AGE_MODEL_CAFFE = os.path.join(MODEL_DIR, 'age_net.caffemodel')
GENDER_MODEL_PROTO = os.path.join(MODEL_DIR, 'deploy_gender.prototxt')
GENDER_MODEL_CAFFE = os.path.join(MODEL_DIR, 'gender_net.caffemodel')
EMOTION_MODEL_PROTO = os.path.join(MODEL_DIR, 'deploy_emotion.prototxt')
EMOTION_MODEL_CAFFE = os.path.join(MODEL_DIR, 'emotion_net.caffemodel')
LANDMARKS_MODEL_PATH = os.path.join(MODEL_DIR, 'shape_predictor_68_face_landmarks.dat')
predictor = None

EYE_AR_THRESH = 0.3 
EYE_AR_CONSEC_FRAMES = 3
blink_counters = {} 
global_blink_counter = 0
global_blink_total = 0
global_last_ear = 0.0

def eye_aspect_ratio(eye):
    A = dist.euclidean(eye[1], eye[5])
    B = dist.euclidean(eye[2], eye[4])
    C = dist.euclidean(eye[0], eye[3])
    ear = (A + B) / (2.0 * C)
    return ear

age_net = None
gender_net = None
emotion_net = None

AGE_BUCKETS = ["(0-2)", "(4-6)", "(8-12)", "(15-18)", "(18-22)", "(22-25)", "(25-32)", "(38-43)", "(48-53)", "(60-100)"]
GENDER_LIST = ['Male', 'Female']
EMOTION_LIST = ['Angry', 'Disgust', 'Fear', 'Happy', 'Sad', 'Surprise', 'Neutral']

try:
    if os.path.exists(LANDMARKS_MODEL_PATH):
        predictor = dlib.shape_predictor(LANDMARKS_MODEL_PATH)
        print("Dlib shape predictor loaded.")
    else:
        print(f"Dlib shape predictor not found at {LANDMARKS_MODEL_PATH}. Liveness detection (blinking) will not work.")
        predictor = None
        
    if os.path.exists(AGE_MODEL_CAFFE) and os.path.exists(AGE_MODEL_PROTO):
        age_net = cv2.dnn.readNet(AGE_MODEL_CAFFE, AGE_MODEL_PROTO)
        print("Age detection model loaded.")
    else:
        print(f"Age model files not found at {AGE_MODEL_CAFFE}, {AGE_MODEL_PROTO}. Age detection will not work.")

    if os.path.exists(GENDER_MODEL_CAFFE) and os.path.exists(GENDER_MODEL_PROTO):
        gender_net = cv2.dnn.readNet(GENDER_MODEL_CAFFE, GENDER_MODEL_PROTO)
        print("Gender detection model loaded.")
    else:
        print(f"Gender model files not found at {GENDER_MODEL_CAFFE}, {GENDER_MODEL_PROTO}. Gender detection will not work.")
    
    if os.path.exists(EMOTION_MODEL_CAFFE) and os.path.exists(EMOTION_MODEL_PROTO):
        emotion_net = cv2.dnn.readNet(EMOTION_MODEL_CAFFE, EMOTION_MODEL_PROTO)
        print("Emotion detection model loaded.")
    else:
        print("Emotion model files not found. Emotion detection will not work.")

except Exception as e:
    print(f"Error loading DNN models or Dlib predictor: {e}. Some features will not work.")
    age_net, gender_net, emotion_net, predictor = None, None, None, None
    
# Define your local timezone

LOCAL_TIMEZONE = pytz.timezone('Asia/Kolkata')
conn = sqlite3.connect('attendance.db')
cursor = conn.cursor()
current_local_time = datetime.now(LOCAL_TIMEZONE).strftime("%Y-%m-%d %H:%M:%S")
print(f"Time to insert: {current_local_time}")
cursor.execute("INSERT INTO attendance_logs (student_roll_no, student_name, period_name, prof_name, timestamp) VALUES (?, ?, ?, ?, ?)", ('TEST', 'Test Student', 'Test Period', 'Test Prof', current_local_time))
conn.commit()
conn.close()
  
# --- User Roles ---
ADMIN_ROLE = 'admin'
PROFESSOR_ROLE = 'professor'

def db_connect():
    """Establishes a connection to the SQLite database and sets row_factory for dictionary-like access."""
    conn = sqlite3.connect(DB_FILE, timeout=10)
    conn.row_factory = sqlite3.Row
    return conn

def get_db():
    db = getattr(g, '_database', None)
    if db is None:
        db = g._database = sqlite3.connect(DB_FILE, timeout=10)
        db.row_factory = sqlite3.Row
    return db

@app.teardown_appcontext
def close_connection(exception):
    db = getattr(g, '_database', None)
    if db is not None:
        db.close()

def init_db():
    conn = db_connect()
    cursor = conn.cursor()

    # Users Table
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL,
        password TEXT NOT NULL,
        role TEXT NOT NULL CHECK (role IN ('admin', 'professor'))
    );
    """)

    # Create default admin and professor users if no users exist
    try:
        cursor.execute("SELECT COUNT(*) FROM users")
        if cursor.fetchone()[0] == 0:
            cursor.execute("INSERT INTO users (username, password, role) VALUES (?, ?, ?)",
                           ('admin', generate_password_hash('adminpass'), 'admin'))
            cursor.execute("INSERT INTO users (username, password, role) VALUES (?, ?, ?)",
                           ('professor1', generate_password_hash('profpass'), 'professor'))
            conn.commit()
            print("Default admin and professor users created.")
    except Exception as e:
        print(f"Error creating default users: {e}")

    # Students Table (ensure gender and age are in the CREATE TABLE too, for fresh installs)
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS students (
        roll_no TEXT PRIMARY KEY,
        name TEXT NOT NULL,
        branch TEXT,
        semester INTEGER,
        admission_year INTEGER,
        subject TEXT,
        face_encoding BLOB,
        gender TEXT,   -- Ensure this is here
        age INTEGER    -- Ensure this is here
    );
    """)

    # Add gender column if it doesn't exist (for existing databases that weren't deleted)
    try:
        cursor.execute("ALTER TABLE students ADD COLUMN gender TEXT;")
        print("Added 'gender' column to 'students' table.")
    except sqlite3.OperationalError as e:
        if "duplicate column name" not in str(e).lower():
            print(f"Error adding 'gender' column to 'students' table: {e}")

    # Add age column if it doesn't exist (for existing databases that weren't deleted)
    try:
        cursor.execute("ALTER TABLE students ADD COLUMN age INTEGER;")
        print("Added 'age' column to 'students' table.")
    except sqlite3.OperationalError as e:
        if "duplicate column name" not in str(e).lower():
            print(f"Error adding 'age' column to 'students' table: {e}")

    # Professors Table
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS professors (
        prof_id TEXT PRIMARY KEY,
        name TEXT NOT NULL,
        department TEXT,
        email TEXT,
        mobile TEXT,
        qualification TEXT,
        experience TEXT,
        achievements TEXT,
        others TEXT,
        photo_data TEXT
    );
    """)

    # Schedule Table
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS schedule (
        period_id INTEGER PRIMARY KEY AUTOINCREMENT,
        period_name TEXT NOT NULL,
        start_time TEXT NOT NULL,
        end_time TEXT NOT NULL,
        prof_id TEXT,
        prof_name TEXT,
        description TEXT,
        branch TEXT,
        semester INTEGER,
        FOREIGN KEY (prof_id) REFERENCES professors (prof_id)
    );
    """)

    try: # Add semester column to schedule if missing
        cursor.execute("ALTER TABLE schedule ADD COLUMN semester INTEGER;")
        print("Added 'semester' column to 'schedule' table.")
    except sqlite3.OperationalError as e:
        if "duplicate column name" not in str(e).lower():
            print(f"Error adding 'semester' column to 'schedule' table: {e}")


    # Attendance Logs Table
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS attendance_logs (
        log_id INTEGER PRIMARY KEY AUTOINCREMENT,
        student_roll_no TEXT NOT NULL,
        student_name TEXT,
        period_name TEXT NOT NULL,
        prof_name TEXT,
        timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (student_roll_no) REFERENCES students (roll_no)
    );
    """)

    conn.commit()
    conn.close()
    print("Database tables created/verified successfully.")
    
def save_face_encodings():
    """Saves the in-memory known_face_encodings and roll_numbers to a pickle file."""
    global known_face_encodings, known_face_roll_numbers
    try:
        with open(ENCODINGS_PATH, 'wb') as f:
            pickle.dump({'encodings': known_face_encodings, 'roll_numbers': known_face_roll_numbers}, f)
        print("Face encodings saved to file.")
    except Exception as e:
        print(f"Error saving encodings file: {e}")

def load_face_encodings():
    """Loads known face encodings and their associated roll numbers from a pickle file into memory."""
    global known_face_encodings, known_face_roll_numbers
    print("Loading known face encodings for live recognition...")
    if os.path.exists(ENCODINGS_PATH):
        try:
            with open(ENCODINGS_PATH, 'rb') as f:
                data = pickle.load(f)
            known_face_encodings = data['encodings']
            known_face_roll_numbers = [str(rn).strip() for rn in data['roll_numbers']]
            print(f"Loaded known_face_roll_numbers: {known_face_roll_numbers}")
            print(f"Loaded {len(known_face_encodings)} known faces.")
        except Exception as e:
            print(f"Error loading encodings file: {e}")
            known_face_encodings, known_face_roll_numbers = [], []
    else:
        print(f"Encodings file not found at {ENCODINGS_PATH}. No faces will be recognized.")
        known_face_encodings, known_face_roll_numbers = [], []

# This block ensures database tables are set up and face encodings are loaded when the app starts
with app.app_context():
    init_db()
    load_face_encodings()

def process_and_encode(image_stream):
    """
    Reads an image from a file-like object, detects a face, and returns its encoding.
    Returns None if no face is found or encoding fails.
    """
    try:
        # Read image bytes and convert to numpy array
        image_bytes = image_stream.read()
        np_arr = np.frombuffer(image_bytes, np.uint8)
        img = cv2.imdecode(np_arr, cv2.IMREAD_COLOR)
        if img is None:
            print("DEBUG: Could not decode image (img is None).")
            flash("Could not decode the uploaded image. Please try a different file.", "danger")
            return None

        # Convert BGR (OpenCV) to RGB (face_recognition)
        rgb_img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        face_locations = face_recognition.face_locations(rgb_img)
        print(f"DEBUG: face_locations found: {face_locations}")
        if not face_locations:
            flash("No face detected in the uploaded image. Please upload a clear photo with a visible face.", "danger")
            return None

        face_encodings = face_recognition.face_encodings(rgb_img, face_locations)
        if not face_encodings:
            flash("Could not extract face encoding. Please try a different photo.", "danger")
            return None

        return face_encodings[0]
    except Exception as e:
        print(f"DEBUG: Exception in process_and_encode: {e}")
        flash(f"Error processing image for face encoding: {e}", "danger")
        return None

# --- Authentication Decorators ---
def login_required(f):
    """Decorator to protect routes requiring a logged-in user."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'logged_in' not in session or not session['logged_in']:
            flash('Please log in to access this page.', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def role_required(roles):
    """Decorator to protect routes based on user roles."""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'logged_in' not in session or not session['logged_in']:
                flash('Please log in to access this page.', 'warning')
                return redirect(url_for('login'))
            
            if 'role' not in session or session['role'] not in roles:
                flash(f'Access denied. You do not have the required role ({", ".join(roles)}).', 'danger')
                if session.get('role') == 'admin':
                    return redirect(url_for('admin_dashboard'))
                elif session.get('role') == 'professor':
                    return redirect(url_for('professor_dashboard'))
                return redirect(url_for('login'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

# --- Helper Functions for Routes ---
def get_distinct_from_db(column_name, table_name='students'):
    """Helper function to fetch unique values for filter dropdowns from a specified table."""
    try:
        with db_connect() as conn:
            cursor = conn.cursor()
            if column_name in ['semester', 'admission_year']:
                query = f"SELECT DISTINCT {column_name} FROM {table_name} WHERE {column_name} IS NOT NULL ORDER BY CAST({column_name} AS INTEGER)"
            else:
                query = f"SELECT DISTINCT {column_name} FROM {table_name} WHERE {column_name} IS NOT NULL ORDER BY {column_name}"
            cursor.execute(query)
            return [item[0] for item in cursor.fetchall()]
    except sqlite3.OperationalError as e:
        print(f"DB Error fetching distinct {column_name} from {table_name}: {e}.")
        return []
    except Exception as e:
        print(f"An unexpected error occurred while fetching distinct {column_name} from {table_name}: {e}")
        return []

def image_to_base64(image_stream):
    """
    Reads an image from a file-like object and converts it to a base64 string.
    Returns None if image cannot be processed.
    """
    try:
        image_bytes = image_stream.read()
        # Optionally, you might want to resize or compress the image here
        # to save database space, especially for larger images.
        # For now, we'll just encode directly.
        return base64.b64encode(image_bytes).decode('utf-8')
    except Exception as e:
        print(f"Error converting image to base64: {e}")
        return None

@app.context_processor
def inject_now():
    return {'now': datetime.now()}

# --- NEW: Landing Page Route ---
@app.route('/landing')
def landing():
    """Renders the landing page."""
    return render_template('landing.html')

# --- Authentication Routes ---
@app.route('/')
def index():
    """Redirects to the landing page or the appropriate dashboard based on session."""
    if 'logged_in' in session and session['logged_in']:
        if session['role'] == 'admin':
            return redirect(url_for('admin_dashboard'))
        elif session['role'] == 'professor':
            return redirect(url_for('professor_dashboard'))
    # If not logged in, go to the landing page instead of directly to login
    return redirect(url_for('landing'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Handles user login."""
    if 'logged_in' in session and session['logged_in']:
        if session['role'] == 'admin':
            return redirect(url_for('admin_dashboard'))
        elif session['role'] == 'professor':
            return redirect(url_for('professor_dashboard'))

    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        conn = db_connect()
        user = conn.execute("SELECT * FROM users WHERE username = ?", (username,)).fetchone()
        conn.close()

        if user and check_password_hash(user['password'], password):
            session['logged_in'] = True
            session['username'] = user['username']
            session['role'] = user['role']
            flash(f'Logged in as {user["username"]} ({user["role"]}).', 'success')
            if user['role'] == 'admin':
                return redirect(url_for('admin_dashboard'))
            elif user['role'] == 'professor':
                return redirect(url_for('professor_dashboard'))
        else:
            flash('Invalid username or password.', 'danger')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    """Logs out the current user."""
    session.pop('logged_in', None)
    session.pop('username', None)
    session.pop('role', None)
    flash('You have been logged out.', 'info')
    return redirect(url_for('login'))

# --- DASHBOARD ROUTES (ROLE-SPECIFIC) ---
@app.route('/dashboard')
@login_required
def dashboard():
    """Redirects to the specific admin or professor dashboard."""
    if session.get('role') == 'admin':
        return redirect(url_for('admin_dashboard'))
    elif session.get('role') == 'professor':
        return redirect(url_for('professor_dashboard'))
    else:
        flash("Unauthorized access. Your role is not recognized.", "danger")
        return redirect(url_for('login'))


@app.route('/admin_dashboard')
@login_required
@role_required(roles=['admin'])
def admin_dashboard():
    """Renders the admin dashboard with summary statistics."""
    conn = db_connect()
    total_students = conn.execute("SELECT COUNT(*) FROM students").fetchone()[0]
    total_professors = conn.execute("SELECT COUNT(*) FROM professors").fetchone()[0]
    total_periods = conn.execute("SELECT COUNT(*) FROM schedule").fetchone()[0]
    local_today_str = datetime.now(LOCAL_TIMEZONE).strftime('%Y-%m-%d')
    today_attendance = conn.execute("SELECT COUNT(*) FROM attendance_logs WHERE date(timestamp) = ?", (local_today_str,)).fetchone()[0]
    conn.close()
    
    return render_template('admin_dashboard.html', 
        total_students=total_students, 
        total_professors=total_professors, 
        total_periods=total_periods,
        today_attendance=today_attendance
    )

@app.route('/professor_dashboard')
@login_required
@role_required(roles=['professor', 'admin'])
def professor_dashboard():
    """Renders the professor dashboard."""
    return render_template('professor_dashboard.html')


# --- ADMIN-ONLY ROUTES ---

@app.route('/students')
@login_required
@role_required(roles=['admin'])
def view_students():
    """Displays a list of all registered students with filtering options."""
    conn = db_connect()

    filter_branch = request.args.get('branch')
    filter_semester = request.args.get('semester')

    students_query = "SELECT roll_no, name, branch, semester, admission_year FROM students WHERE 1=1"
    student_query_params = []

    if filter_branch and filter_branch != 'all':
        students_query += " AND branch = ?"
        student_query_params.append(filter_branch)
    
    if filter_semester and filter_semester != 'all':
        students_query += " AND semester = ?"
        student_query_params.append(filter_semester)

    students_query += " ORDER BY name"
    
    students_list = conn.execute(students_query, student_query_params).fetchall()

    available_branches = get_distinct_from_db('branch', 'students')
    available_semesters = get_distinct_from_db('semester', 'students')
    
    conn.close()
    return render_template('students.html', 
                           students=students_list, 
                           available_branches=available_branches, 
                           available_semesters=available_semesters,
                           selected_branch=filter_branch, 
                           selected_semester=filter_semester)

@app.route('/add_student', methods=['GET', 'POST'])
@login_required
@role_required(roles=['admin'])
def add_student():
    current_year = datetime.now().year

    if request.method == 'POST':
        roll_no = request.form['roll_no'].strip()
        name = request.form['name'].strip()
        branch = request.form.get('branch', '').strip()
        semester = request.form.get('semester')
        admission_year = request.form.get('admission_year')
        subject = request.form.get('subject', '').strip()
        
        photo_upload_method = request.form.get('photo_upload_method')

        if not all([roll_no, name, branch, semester, admission_year]):
            flash("All fields (Roll Number, Name, Branch, Semester, Year of Admission) are required.", "danger")
            return render_template('add_student.html', current_year=current_year, **request.form)

        face_encoding = None
        if photo_upload_method == 'file_upload':
            photo_file = request.files.get('photo_file')
            if photo_file and photo_file.filename != '':
                face_encoding = process_and_encode(photo_file.stream)
            else:
                flash("No photo file was uploaded. Please select a photo file.", "warning")
                return render_template('add_student.html', current_year=current_year, **request.form)
        elif photo_upload_method == 'camera_capture':
            photo_data = request.form.get('photo_data')
            if photo_data:
                header, base64_string = photo_data.split(',', 1) 
                image_bytes = base64.b64decode(base64_string)
                face_encoding = process_and_encode(BytesIO(image_bytes))
            else:
                flash("No photo was captured from the camera. Please capture a photo.", "warning")
                return render_template('add_student.html', current_year=current_year, **request.form)
        else:
            flash("No photo upload method selected.", "danger")
            return render_template('add_student.html', current_year=current_year, **request.form)
            
        if face_encoding is None:
            return render_template('add_student.html', current_year=current_year, **request.form)

        try:
            encoding_blob = pickle.dumps(face_encoding)

            with db_connect() as conn:
                conn.execute("""
                    INSERT INTO students (roll_no, name, branch, semester, admission_year, subject, face_encoding)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (roll_no, name, branch, semester, admission_year, subject, encoding_blob))
                conn.commit()
            
            flash(f"Student '{name}' ({roll_no}) and their face encoding have been added successfully!", "success")
            load_face_encodings()
            return redirect(url_for('view_students'))

        except sqlite3.IntegrityError:
            flash(f"A student with Roll Number '{roll_no}' already exists. Please use a unique Roll Number.", "danger")
        except Exception as e:
            flash(f"A database error occurred while adding student: {e}", "danger")
            
        return render_template('add_student.html', current_year=current_year, **request.form)

    return render_template('add_student.html', current_year=current_year)


@app.route('/edit_student/<roll_no>', methods=['GET', 'POST'])
@login_required
@role_required(roles=['admin'])
def edit_student(roll_no):
    conn = db_connect()
    student = conn.execute("SELECT roll_no, name, branch, semester, admission_year, subject, face_encoding FROM students WHERE roll_no = ?", (roll_no,)).fetchone()
    conn.close()

    if not student:
        flash(f"Student with Roll Number '{roll_no}' not found.", "danger")
        return redirect(url_for('view_students'))

    current_year = datetime.now().year

    if request.method == 'POST':
        submitted_roll_no = request.form['roll_no'].strip()
        if submitted_roll_no != roll_no:
            flash("Attempted to change Roll Number, which is not allowed.", "danger")
            return render_template('edit_student.html', student=student, current_year=current_year, **request.form)

        name = request.form['name'].strip()
        branch = request.form.get('branch', '').strip()
        semester = request.form.get('semester')
        admission_year = request.form.get('admission_year')
        subject = request.form.get('subject', '').strip()

        if not all([submitted_roll_no, name, branch, semester, admission_year]):
            flash("All required fields (Roll Number, Name, Branch, Semester, Year of Admission) are necessary for update.", "danger")
            return render_template('edit_student.html', student=student, current_year=current_year, **request.form)

        photo_upload_method = request.form.get('photo_upload_method')
        new_face_encoding_blob = None
        face_encoding_updated = False

        if photo_upload_method == 'file_upload':
            photo_file = request.files.get('photo_file')
            if photo_file and photo_file.filename != '':
                new_face_encoding = process_and_encode(photo_file.stream)
                if new_face_encoding is None:
                    return render_template('edit_student.html', student=student, current_year=current_year, **request.form)
                new_face_encoding_blob = pickle.dumps(new_face_encoding)
                face_encoding_updated = True
            elif student['face_encoding']:
                new_face_encoding_blob = student['face_encoding']
                face_encoding_updated = False
            else:
                flash("No new photo file was uploaded and no existing face encoding is present. A photo is required for recognition.", "danger")
                return render_template('edit_student.html', student=student, current_year=current_year, **request.form)

        elif photo_upload_method == 'camera_capture':
            photo_data = request.form.get('photo_data')
            if photo_data:
                header, base64_string = photo_data.split(',', 1)
                image_bytes = base64.b64decode(base64_string)
                new_face_encoding = process_and_encode(BytesIO(image_bytes))
                if new_face_encoding is None:
                    return render_template('edit_student.html', student=student, current_year=current_year, **request.form)
                new_face_encoding_blob = pickle.dumps(new_face_encoding)
                face_encoding_updated = True
            elif student['face_encoding']:
                new_face_encoding_blob = student['face_encoding']
                face_encoding_updated = False
            else:
                flash("No photo was captured and no existing face encoding is present. A photo is required for recognition.", "danger")
                return render_template('edit_student.html', student=student, current_year=current_year, **request.form)

        elif photo_upload_method == 'keep_current':
            if student['face_encoding']:
                new_face_encoding_blob = student['face_encoding']
                face_encoding_updated = False
            else:
                flash("Cannot keep current photo: no existing face encoding found for this student. Please upload or capture a new photo.", "danger")
                return render_template('edit_student.html', student=student, current_year=current_year, **request.form)
        else:
            flash("Invalid photo upload method selected.", "danger")
            return render_template('edit_student.html', student=student, current_year=current_year, **request.form)

        try:
            conn = db_connect()
            conn.execute("""
                UPDATE students SET 
                    name = ?, branch = ?, semester = ?, admission_year = ?, subject = ?, face_encoding = ?
                WHERE roll_no = ?
            """, (name, branch, semester, admission_year, subject, new_face_encoding_blob, roll_no))
            conn.commit()
            conn.close()

            flash(f"Student '{name}' ({roll_no}) details updated successfully!", "success")
            
            if face_encoding_updated:
                load_face_encodings()
                flash("Face encodings reloaded for live recognition.", "info")

            return redirect(url_for('view_students'))

        except Exception as e:
            flash(f"A database error occurred while updating student: {e}", "danger")
            return render_template('edit_student.html', student=student, current_year=current_year, **request.form)
            
    return render_template('edit_student.html', student=student, current_year=current_year)


@app.route('/delete_student/<roll_no>', methods=['POST'])
@login_required
@role_required(roles=['admin'])
def delete_student(roll_no):
    conn = db_connect()
    try:
        attendance_count = conn.execute("SELECT COUNT(*) FROM attendance_logs WHERE student_roll_no = ?", (roll_no,)).fetchone()[0]
        if attendance_count > 0:
            flash(f"Cannot delete student {roll_no}. They have {attendance_count} attendance records. Delete attendance records first or manually update them.", "danger")
            return redirect(url_for('view_students'))
        
        cursor = conn.cursor()
        cursor.execute("DELETE FROM students WHERE roll_no = ?", (roll_no,))
        conn.commit()

        global known_face_encodings, known_face_roll_numbers
        if roll_no in known_face_roll_numbers:
            idx = known_face_roll_numbers.index(roll_no)
            known_face_roll_numbers.pop(idx)
            known_face_encodings.pop(idx)
            save_face_encodings()
            flash(f"Student {roll_no} and their face encoding deleted successfully! Server restart recommended for live attendance.", "success")
        else:
            flash(f"Student {roll_no} deleted successfully! (No associated face encoding found).", "success")
        
    except Exception as e:
        flash(f"Error deleting student {roll_no}: {e}", "danger")
    finally:
        conn.close()
    return redirect(url_for('view_students'))


@app.route('/professors')
@login_required
@role_required(roles=['admin'])
def view_professors():
    """Displays a list of all registered professors."""
    conn = db_connect()
    prof_list = conn.execute("SELECT * FROM professors ORDER BY name").fetchall()
    conn.close()
    return render_template('professors.html', professors=prof_list)

@app.route('/add_professor', methods=['GET', 'POST'])
@login_required
@role_required(roles=['admin'])
def add_professor():
    """Handles adding a new professor, including their photo (base64)."""
    if request.method == 'POST':
        prof_id = request.form['prof_id'].strip()
        name = request.form['name'].strip()
        department = request.form['department'].strip()
        email = request.form.get('email', '').strip()
        mobile = request.form.get('mobile', '').strip()
        qualification = request.form.get('qualification', '').strip()
        experience = request.form.get('experience', '').strip()
        achievements = request.form.get('achievements', '').strip()
        others = request.form.get('others', '').strip()
        
        photo_upload_method = request.form.get('photo_upload_method') # New: Get method
        photo_data_b64 = None # Will store the base64 photo string

        # Handle photo upload based on method
        if photo_upload_method == 'file_upload':
            photo_file = request.files.get('photo_file')
            if photo_file and photo_file.filename != '':
                photo_data_b64 = image_to_base64(photo_file.stream)
                if photo_data_b64 is None:
                    flash("Could not process the uploaded photo file. Please try a different file.", "danger")
                    return render_template('add_professor.html', **request.form)
            else:
                flash("No photo file was uploaded. Please select a photo file or use the camera.", "warning")
                return render_template('add_professor.html', **request.form)
        elif photo_upload_method == 'camera_capture':
            captured_photo_data = request.form.get('photo_data_capture') # New name for camera data
            if captured_photo_data:
                # Assuming captured_photo_data is already a base64 data URL (e.g., "data:image/png;base64,...")
                # We need to strip the header: "data:image/png;base64,"
                if ',' in captured_photo_data:
                    header, base64_string = captured_photo_data.split(',', 1)
                    photo_data_b64 = base64_string
                else:
                    photo_data_b64 = captured_photo_data # If no header, assume raw base64
                
                if photo_data_b64 is None or photo_data_b64 == "":
                     flash("Could not process captured photo. Please try again.", "danger")
                     return render_template('add_professor.html', **request.form)
            else:
                flash("No photo was captured from the camera. Please capture a photo.", "warning")
                return render_template('add_professor.html', **request.form)
        else:
            flash("No photo upload method selected. Please select one.", "danger")
            return render_template('add_professor.html', **request.form)

        if not all((prof_id, name, department)):
            flash("Professor ID, Name, and Department are required.", "danger")
            return render_template('add_professor.html', **request.form)
            
        try:
            conn = db_connect()
            conn.execute("""
                INSERT INTO professors 
                (prof_id, name, department, email, mobile, qualification, experience, achievements, others, photo_data) 
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (prof_id, name, department, email, mobile, qualification, experience, achievements, others, photo_data_b64)) # Use photo_data_b64
            conn.commit()
            conn.close()
            flash(f"Professor '{name}' added successfully!", "success")
            return redirect(url_for('view_professors'))
        except sqlite3.IntegrityError:
            flash(f"A professor with ID '{prof_id}' already exists. Please use a unique Professor ID.", "danger")
            return render_template('add_professor.html', **request.form)
        except Exception as e:
            flash(f"An unexpected error occurred while adding professor: {e}", "danger")
            return render_template('add_professor.html', **request.form)
    
    return render_template('add_professor.html')


@app.route('/edit_professor_profile/<prof_id>', methods=['GET', 'POST'])
@login_required
@role_required(roles=['admin'])
def edit_professor_profile(prof_id):
    conn = db_connect()
    professor = conn.execute("SELECT * FROM professors WHERE prof_id = ?", (prof_id,)).fetchone()
    conn.close() # Close initial connection

    if not professor:
        flash(f"Professor with ID {prof_id} not found.", "danger")
        return redirect(url_for('view_professors'))

    if request.method == 'POST':
        name = request.form['name'].strip()
        department = request.form['department'].strip()
        email = request.form.get('email', '').strip()
        mobile = request.form.get('mobile', '').strip()
        qualification = request.form.get('qualification', '').strip()
        experience = request.form.get('experience', '').strip()
        achievements = request.form.get('achievements', '').strip()
        others = request.form.get('others', '').strip()
        
        photo_upload_method = request.form.get('photo_upload_method')
        photo_data_to_save = None

        if photo_upload_method == 'file_upload':
            photo_file = request.files.get('photo_file')
            if photo_file and photo_file.filename != '':
                photo_data_to_save = image_to_base64(photo_file.stream)
                if photo_data_to_save is None:
                    flash("Could not process the uploaded photo file. Please try a different file.", "danger")
                    return render_template('edit_professor_profile.html', professor=professor, **request.form)
            else: # No new file uploaded, keep existing if present
                photo_data_to_save = professor['photo_data']
        elif photo_upload_method == 'camera_capture':
            captured_photo_data = request.form.get('photo_data_capture')
            if captured_photo_data:
                if ',' in captured_photo_data:
                    header, base64_string = captured_photo_data.split(',', 1)
                    photo_data_to_save = base64_string
                else:
                    photo_data_to_save = captured_photo_data
                
                if photo_data_to_save is None or photo_data_to_save == "":
                     flash("Could not process captured photo. Please try again.", "danger")
                     return render_template('edit_professor_profile.html', professor=professor, **request.form)
            else: # No new photo captured, keep existing if present
                photo_data_to_save = professor['photo_data']
        elif photo_upload_method == 'keep_current':
            photo_data_to_save = professor['photo_data']
        elif photo_upload_method == 'clear_photo': # New option to clear the photo
            photo_data_to_save = None
        else:
            flash("Invalid photo upload method selected.", "danger")
            return render_template('edit_professor_profile.html', professor=professor, **request.form)

        try:
            conn = db_connect() # Re-open connection for update
            conn.execute("""
                UPDATE professors SET 
                    name = ?, department = ?, email = ?, mobile = ?, 
                    qualification = ?, experience = ?, achievements = ?, others = ?, photo_data = ?
                WHERE prof_id = ?
            """, (name, department, email, mobile, qualification, experience, achievements, others, photo_data_to_save, prof_id))
            conn.commit()
            conn.close()
            flash(f"Professor '{name}' profile updated successfully!", "success")
            return redirect(url_for('view_professor_profile', prof_id=prof_id))
        except Exception as e:
            flash(f"Error updating professor profile: {e}", "danger")
            # Pass professor object back in case of error
            updated_professor = conn.execute("SELECT * FROM professors WHERE prof_id = ?", (prof_id,)).fetchone()
            conn.close()
            return render_template('edit_professor_profile.html', professor=updated_professor, **request.form)
    
    # For GET request, ensure professor data is passed
    return render_template('edit_professor_profile.html', professor=professor)
@app.route('/delete_professor/<prof_id>', methods=['POST'])
@login_required
@role_required(roles=['admin'])
def delete_professor(prof_id):
    conn = db_connect()
    try:
        schedule_count = conn.execute("SELECT COUNT(*) FROM schedule WHERE prof_id = ?", (prof_id,)).fetchone()[0]
        if schedule_count > 0:
            flash(f"Cannot delete professor {prof_id}. They are assigned to {schedule_count} periods. Please reassign or delete these periods first.", "danger")
            return redirect(url_for('view_professors'))

        attendance_count = conn.execute("SELECT COUNT(*) FROM attendance_logs WHERE prof_name = (SELECT name FROM professors WHERE prof_id = ?)", (prof_id,)).fetchone()[0]
        if attendance_count > 0:
            flash(f"Cannot delete professor {prof_id}. They are referenced in {attendance_count} attendance records. Please delete these attendance records first.", "danger")
            return redirect(url_for('view_professors'))

        cursor = conn.cursor()
        cursor.execute("DELETE FROM professors WHERE prof_id = ?", (prof_id,))
        conn.commit()
        flash(f"Professor {prof_id} deleted successfully!", "success")
    except Exception as e:
        flash(f"Error deleting professor {prof_id}: {e}", "danger")
    finally:
        conn.close()
    return redirect(url_for('view_professors'))

@app.route('/professor_profile/<prof_id>', methods=['GET'])
@login_required
@role_required(roles=['admin'])
def view_professor_profile(prof_id):
    conn = db_connect()
    professor = conn.execute("SELECT * FROM professors WHERE prof_id = ?", (prof_id,)).fetchone()
    conn.close()
    if not professor:
        flash(f"Professor with ID {prof_id} not found.", "danger")
        return redirect(url_for('view_professors'))
    
    return render_template('professor_profile.html', professor=professor)


@app.route('/schedule', methods=['GET', 'POST'])
@login_required
@role_required(roles=['admin', 'professor'])
def manage_schedule():
    conn = db_connect()
    if request.method == 'POST':
        period_name = request.form['period_name'].strip()
        start_time = request.form['start_time'].strip()
        end_time = request.form['end_time'].strip()
        prof_id = request.form['prof_id']
        period_description = request.form.get('description', '').strip()
        
        # --- NEW: Get Branch and Semester from form ---
        branch = request.form.get('branch', '').strip()
        semester = request.form.get('semester')
        try:
            semester = int(semester) if semester else None # Convert to int, handle empty
        except ValueError:
            flash("Semester must be a valid number.", "danger")
            # If there's an error, re-render form with current data
            schedule_list = conn.execute("SELECT * FROM schedule ORDER BY start_time").fetchall()
            prof_list = conn.execute("SELECT prof_id, name FROM professors ORDER BY name").fetchall()
            available_branches = get_distinct_from_db('branch', 'students')
            available_semesters = get_distinct_from_db('semester', 'students')
            conn.close()
            return render_template('schedule.html', 
                                   schedule=schedule_list, 
                                   professors=prof_list, 
                                   available_branches=available_branches,
                                   available_semesters=available_semesters,
                                   # Pass back current form values in case of error
                                   form_data=request.form)


        prof_record = conn.execute("SELECT name FROM professors WHERE prof_id = ?", (prof_id,)).fetchone()
        prof_name = prof_record['name'] if prof_record else "N/A"
        
        try:
            conn.execute("""
                INSERT INTO schedule (period_name, start_time, end_time, prof_id, prof_name, description, branch, semester)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (period_name, start_time, end_time, prof_id, prof_name, period_description, branch, semester)) # Added branch, semester
            conn.commit()
            conn.close()
            flash(f"Period '{period_name}' added to the schedule.", "success")
            return redirect(url_for('manage_schedule'))
        except Exception as e:
            flash(f"Error adding period: {e}", "danger")
            # Re-fetch data for rendering in case of DB error
            schedule_list = conn.execute("SELECT * FROM schedule ORDER BY start_time").fetchall()
            prof_list = conn.execute("SELECT prof_id, name FROM professors ORDER BY name").fetchall()
            available_branches = get_distinct_from_db('branch', 'students')
            available_semesters = get_distinct_from_db('semester', 'students')
            conn.close()
            return render_template('schedule.html', 
                                   schedule=schedule_list, 
                                   professors=prof_list, 
                                   available_branches=available_branches,
                                   available_semesters=available_semesters,
                                   # Pass back current form values in case of error
                                   form_data=request.form)


    # --- For GET Request (and after successful POST redirect) ---
    schedule_list = conn.execute("SELECT * FROM schedule ORDER BY start_time").fetchall()
    prof_list = conn.execute("SELECT prof_id, name FROM professors ORDER BY name").fetchall()
    
    # --- NEW: Fetch available branches and semesters from students data ---
    available_branches = get_distinct_from_db('branch', 'students')
    available_semesters = get_distinct_from_db('semester', 'students')

    conn.close()
    return render_template('schedule.html', 
                           schedule=schedule_list, 
                           professors=prof_list, 
                           available_branches=available_branches, # Pass to template
                           available_semesters=available_semesters, # Pass to template
                           form_data={}) # Pass empty dict for initial GET
    
    
@app.route('/delete_period/<int:period_id>', methods=['POST'])
@login_required
@role_required(roles=['admin', 'professor'])
def delete_period(period_id):
    conn = db_connect()
    try:
        period_record = conn.execute("SELECT period_name FROM schedule WHERE period_id = ?", (period_id,)).fetchone()
        period_name = period_record['period_name'] if period_record else f"Period ID {period_id}"

        cursor = conn.cursor()
        cursor.execute("DELETE FROM attendance_logs WHERE period_name = ?", (period_name,))
        
        cursor.execute("DELETE FROM schedule WHERE period_id = ?", (period_id,))
        conn.commit()
        flash(f"Period '{period_name}' and its attendance logs deleted successfully!", "success")
    except Exception as e:
        flash(f"Error deleting period '{period_name}': {e}", "danger")
    finally:
        conn.close()
    return redirect(url_for('manage_schedule'))


def get_attendance_data_for_display(filters):
    conn = get_db()
    
    # 1. Get all students that match branch and semester filters
    all_students_query = "SELECT roll_no, name, branch, semester FROM students WHERE 1=1"
    student_query_params = []
    
    if filters.get('branch') and filters['branch'] != 'all':
        all_students_query += " AND branch = ?"
        student_query_params.append(filters['branch'])
    
    if filters.get('semester') and filters['semester'] != 'all':
        all_students_query += " AND semester = ?"
        try:
            student_query_params.append(int(filters['semester']))
        except (ValueError, TypeError):
            pass
            
    all_eligible_students = conn.execute(all_students_query, student_query_params).fetchall()
    
    present_logs_query = "SELECT log_id, student_roll_no, student_name, period_name, prof_name, timestamp FROM attendance_logs WHERE 1=1"
    present_log_params = []

    if filters.get('period_id') and filters['period_id'] != 'all':
        period_record = conn.execute("SELECT period_name, prof_name FROM schedule WHERE period_id = ?", (filters['period_id'],)).fetchone()
        if period_record:
            present_logs_query += " AND period_name = ?"
            present_log_params.append(period_record['period_name'])
        else:
            present_log_params.append("NON_EXISTENT_PERIOD")
            
    if filters.get('date'):
        present_logs_query += " AND date(timestamp) = ?"
        present_log_params.append(filters['date'])
        
    present_logs_raw = conn.execute(present_logs_query, present_log_params).fetchall()
    
    present_students_map = {}
    for log in present_logs_raw:
        present_students_map[log['student_roll_no']] = log
        
    final_attendance_records = []
    
    period_prof_name_for_absent = 'N/A'
    if filters.get('period_id') and filters['period_id'] != 'all':
        period_record = conn.execute("SELECT prof_name FROM schedule WHERE period_id = ?", (filters['period_id'],)).fetchone()
        if period_record:
            period_prof_name_for_absent = period_record['prof_name']

    for student in all_eligible_students:
        roll_no = student['roll_no']
        
        if roll_no in present_students_map:
            log = present_students_map[roll_no]
            log_dict = dict(log)
            log_dict['status'] = 'Present'
            
            # --- CORRECTED TIMESTAMP CONVERSION ---
            # Inside get_attendance_data_for_display, within the loop where log_dict is processed:
            if log_dict['timestamp']:
                try:
                    # Assume DB stores LOCAL_TIMEZONE directly
                    dt_naive = datetime.strptime(str(log_dict['timestamp']), '%Y-%m-%d %H:%M:%S')
                    # Corrected indentation: These lines need to be inside the try block
                    dt_aware_local = LOCAL_TIMEZONE.localize(dt_naive) # Localize directly
                    log_dict['timestamp'] = dt_aware_local.strftime('%Y-%m-%d %H:%M:%S')
                except ValueError:
                    log_dict['timestamp'] = str(log_dict['timestamp']) # Keep as string if parsing fails
            else:
                log_dict['timestamp'] = 'N/A' # In case timestamp is None
            # --- END CORRECTED TIMESTAMP CONVERSION ---
            
            # Populate branch and semester from student for Present records too
            log_dict['student_branch'] = student['branch'] if student['branch'] else 'N/A'
            log_dict['student_semester'] = student['semester'] if student['semester'] else 'N/A'

            final_attendance_records.append(log_dict)
        else: # This 'else' belongs to 'if roll_no in present_students_map:'
            if (filters.get('period_id') and filters['period_id'] != 'all') and \
               filters.get('date'):
                
                if filters.get('status') == 'absent' or filters.get('status') == 'all':
                    # For Absent records, populate branch and semester from student
                    absent_student_branch = student['branch'] if student['branch'] else 'N/A'
                    absent_student_semester = student['semester'] if student['semester'] else 'N/A'
    
                    final_attendance_records.append({
                        'log_id': None,
                        'student_roll_no': student['roll_no'],
                        'student_name': student['name'],
                        'period_name': conn.execute("SELECT period_name FROM schedule WHERE period_id = ?", (filters['period_id'],)).fetchone()['period_name'] if filters['period_id'] != 'all' else 'N/A',
                        'prof_name': period_prof_name_for_absent,
                        'timestamp': 'ABSENT', # 'ABSENT' string for absent records
                        'status': 'Absent',
                        'student_branch': absent_student_branch, # Populate branch for absent
                        'student_semester': absent_student_semester # Populate semester for absent
                    })
    
    if filters.get('status') and filters['status'] != 'all':
        final_attendance_records = [
            record for record in final_attendance_records if record['status'].lower() == filters['status'].lower()
        ]

    final_attendance_records.sort(key=lambda x: x['student_name'])

    return final_attendance_records

@app.route('/manual_attendance', methods=['GET', 'POST'])
@login_required
@role_required(roles=['admin', 'professor'])
def manual_attendance():
    conn = db_connect()

    filter_branch = request.args.get('branch')
    filter_semester = request.args.get('semester')
    filter_period_id_from_url = request.args.get('period_id')

    period_id_for_marking = request.form.get('period_id')
    
    # --- CORRECTED LINE: Ensure form_data_for_template always behaves like a MultiDict ---
    # For POST requests, use request.form (which is a MultiDict).
    # For GET requests, use request.args (which is an ImmutableMultiDict, also supporting .getlist()).
    form_data_for_template = request.form if request.method == 'POST' else request.args

    if request.method == 'POST':
        action = request.form.get('action')

        if action == 'mark_attendance':
            # request.form.getlist('roll_nos') is correct here as it's a POST request
            selected_roll_nos = request.form.getlist('roll_nos') 
            
            current_filter_branch = request.form.get('current_filter_branch')
            current_filter_semester = request.form.get('current_filter_semester')
            current_filter_period_id = request.form.get('current_filter_period_id')

            if not selected_roll_nos:
                flash("Please select at least one student to mark as present.", "warning")
                conn.close()
                # When redirecting after a POST, pass current form data as query args
                # to re-populate the form and filters correctly.
                return redirect(url_for(
                    'manual_attendance', 
                    branch=current_filter_branch, 
                    semester=current_filter_semester, 
                    period_id=current_filter_period_id, 
                    _external=True,
                    # No need to pass **request.form here if we are passing specific filters
                    # as request.form is only available for the current POST.
                    # The template will use selected_* variables for filters and form_data for checkboxes.
                ))
            
            if not period_id_for_marking or period_id_for_marking == 'all':
                flash("Please select a specific period from the 'Select Period to Mark Attendance For' dropdown.", "warning")
                conn.close()
                return redirect(url_for(
                    'manual_attendance', 
                    branch=current_filter_branch, 
                    semester=current_filter_semester, 
                    period_id=current_filter_period_id,
                    _external=True,
                    # No need for **request.form here either
                ))

            period = conn.execute("SELECT period_name, prof_name FROM schedule WHERE period_id = ?", (period_id_for_marking,)).fetchone()
            if not period:
                flash("Selected period not found in schedule.", "danger")
                conn.close()
                return redirect(url_for(
                    'manual_attendance', 
                    branch=current_filter_branch, 
                    semester=current_filter_semester, 
                    period_id=current_filter_period_id,
                    _external=True,
                    # No need for **request.form here either
                ))

            period_name, prof_name = period['period_name'], period['prof_name']
            logged_count = 0
            
            for roll_no in selected_roll_nos:
                student = conn.execute("SELECT name FROM students WHERE roll_no = ?", (roll_no,)).fetchone()
                if student:
                    student_name = student['name']
                    local_now = datetime.now(LOCAL_TIMEZONE)
                    timestamp_str = local_now.strftime("%Y-%m-%d %H:%M:%S") 
                    today_str = local_now.strftime("%Y-%m-%d")
                    
                    existing_log = conn.execute("SELECT 1 FROM attendance_logs WHERE student_roll_no = ? AND period_name = ? AND date(timestamp) = ?", (roll_no, period_name, today_str)).fetchone()
                    
                    if not existing_log:
                        conn.execute("INSERT INTO attendance_logs (student_roll_no, student_name, period_name, prof_name, timestamp) VALUES (?, ?, ?, ?, ?)",
                                     (roll_no, student_name, period_name, prof_name, timestamp_str))
                        logged_count += 1
                    else:
                        print(f"DEBUG_MANUAL_ATTENDANCE: Attendance already exists for {roll_no}")

            conn.commit()
            conn.close()

            if logged_count > 0:
                flash(f"Successfully marked attendance for {logged_count} student(s) in {period_name}.", "success")
            else:
                flash("No new attendance was logged (students may have already been marked present for this period today).", "info")
            
            # Redirect to the GET route for manual_attendance, preserving filters
            return redirect(url_for(
                'manual_attendance', 
                branch=current_filter_branch, 
                semester=current_filter_semester, 
                period_id=current_filter_period_id, 
                _external=True
            ))
        else:
            print("DEBUG_MANUAL_ATTENDANCE: POST request with no 'mark_attendance' action detected. Ignoring.")
            pass # Keep existing behavior for other POST actions if any
    
    # --- GET request logic (or after a redirect from POST) ---
    students_query = "SELECT roll_no, name, branch, semester FROM students WHERE 1=1"
    student_query_params = []

    if filter_branch and filter_branch != 'all':
        students_query += " AND branch = ?"
        student_query_params.append(filter_branch)
    
    if filter_semester and filter_semester != 'all':
        students_query += " AND semester = ?"
        try:
            student_query_params.append(int(filter_semester))
        except (ValueError, TypeError):
            print(f"WARNING: Invalid semester filter '{filter_semester}' received. Skipping semester filter.")
            filter_semester = None # Clear invalid semester
            # Note: filter_semester might be 'all' or actual value, ensure it's propagated correctly
            # back to template. Already handled by selected_semester.

    students_query += " ORDER BY name"
    print(f"DEBUG_MANUAL_ATTENDANCE: Fetching students with Query: '{students_query}' and Params: {student_query_params}")
    students_list = conn.execute(students_query, student_query_params).fetchall()
    print(f"DEBUG_MANUAL_ATTENDANCE: Fetched {len(students_list)} students for display.")

    available_branches = get_distinct_from_db('branch', 'students')
    available_semesters = get_distinct_from_db('semester', 'students')
    schedule_list = conn.execute("SELECT period_id, period_name, start_time, end_time, branch, semester FROM schedule ORDER BY start_time").fetchall()
    
    conn.close()

    return render_template(
        'manual_attendance.html',
        students=students_list,
        schedule=schedule_list,
        available_branches=available_branches,
        available_semesters=available_semesters,
        selected_branch=filter_branch, # These variables are for displaying current filters
        selected_semester=filter_semester,
        selected_period_id=filter_period_id_from_url,
        form_data=form_data_for_template # Pass the correctly initialized form_data
    )
    
def get_filtered_attendance_data(filters):
    conn = get_db()
    log_entries = []
    
    base_query = "SELECT log_id, student_roll_no, student_name, period_name, prof_name, timestamp FROM attendance_logs WHERE 1=1"
    query_params = []

    if filters.get('branch'):
        base_query += " AND student_roll_no IN (SELECT roll_no FROM students WHERE branch = ?)"
        query_params.append(filters['branch'])
    if filters.get('semester'):
        base_query += " AND student_roll_no IN (SELECT roll_no FROM students WHERE semester = ?)"
        query_params.append(filters['semester'])
    if filters.get('period_name'):
        base_query += " AND period_name = ?"
        query_params.append(filters['period_name'])
    if filters.get('date'):
        base_query += " AND date(timestamp) = ?"
        query_params.append(filters['date'])
    
    base_query += " ORDER BY timestamp DESC"
    present_students_raw = conn.execute(base_query, query_params).fetchall()
    
    present_roll_nos = set()
    for entry in present_students_raw:
        log_dict = dict(entry)
        student_info = conn.execute("SELECT branch, semester FROM students WHERE roll_no = ?", (log_dict['student_roll_no'],)).fetchone()
        log_dict['student_branch'] = student_info['branch'] if student_info else 'N/A'
        log_dict['student_semester'] = student_info['semester'] if student_info else 'N/A'
        log_dict['status'] = 'Present'
        log_entries.append(log_dict)
        present_roll_nos.add(log_dict['student_roll_no'])

    if filters.get('period_name') and filters.get('date'):
        all_students_query = "SELECT roll_no, name, branch, semester FROM students WHERE 1=1"
        student_params = []
        if filters.get('branch'):
            all_students_query += " AND branch = ?"
            student_params.append(filters['branch'])
        if filters.get('semester'):
            all_students_query += " AND semester = ?"
            student_params.append(filters['semester'])
        
        all_eligible_students = conn.execute(all_students_query, student_params).fetchall()
        
        for student in all_eligible_students:
            if student['roll_no'] not in present_roll_nos:
                log_entries.append({
                    'student_roll_no': student['roll_no'],
                    'student_name': student['name'],
                    'period_name': filters['period_name'],
                    'prof_name': 'N/A',
                    'timestamp': filters['date'],
                    'student_branch': student['branch'],
                    'student_semester': student['semester'],
                    'status': 'Absent'
                })
    return log_entries

@app.route('/log', methods=['GET'])
@login_required
@role_required(roles=['admin', 'professor'])
def view_attendance_log():
    # Retrieve filter parameters
    filter_branch = request.args.get('branch')
    filter_semester = request.args.get('semester')
    filter_period_id = request.args.get('period_id')
    filter_date_str = request.args.get('date')
    filter_status = request.args.get('status', 'all') # Default to 'all' status

    # Create the 'filters' dictionary to pass to the template
    current_filters = {
        'branch': filter_branch,
        'semester': filter_semester,
        'period_id': filter_period_id,
        'date': filter_date_str,
        'status': filter_status # Pass the status filter
    }

    # Get attendance data using the new helper function
    log_entries = get_attendance_data_for_display(current_filters)
    
    conn = get_db() # Get a connection for dropdowns
    available_branches = get_distinct_from_db('branch', 'students')
    available_semesters = get_distinct_from_db('semester', 'students')
    schedule_periods = conn.execute("SELECT period_id, period_name, start_time, end_time FROM schedule ORDER BY period_name").fetchall() 
    conn.close() # Close connection after getting dropdown data

    return render_template(
        'log.html', 
        logs=log_entries, 
        schedule_periods=schedule_periods, 
        available_branches=available_branches, 
        available_semesters=available_semesters, 
        filters=current_filters # Pass the 'filters' dictionary
    )
    
@app.route('/delete_log_entry/<int:log_id>', methods=['POST'])
@login_required
@role_required(roles=['admin', 'professor'])
def delete_log_entry(log_id):
    conn = db_connect()
    try:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM attendance_logs WHERE log_id = ?", (log_id,))
        conn.commit()
        flash(f"Attendance log entry {log_id} deleted successfully!", "success")
    except Exception as e:
        flash(f"Error deleting log entry {log_id}: {e}", "danger")
    finally:
        conn.close()
    return redirect(url_for('view_attendance_log'))

# --- Gemini API Routes ---

@app.route('/generate_period_description', methods=['POST'])
@login_required
def generate_period_description():
    data = request.get_json()
    period_name = data.get('period_name')
    professor_name = data.get('professor_name')

    if not period_name or not professor_name:
        return jsonify({'status': 'error', 'message': 'Period name and professor name are required.'}), 400

    prompt = f"Generate a brief, engaging description (approx. 2-3 sentences) for a university course named '{period_name}' taught by Professor '{professor_name}'. Focus on what students will learn or experience, and keep it concise for a course catalog. Avoid phrases like 'This course will...' or 'Students will learn...'. Instead, weave the learning outcomes naturally into the description."

    try:
        chat_history = []
        chat_history.append({"role": "user", "parts": [{"text": prompt}]}) 
        
        payload = {
            "contents": chat_history,
            "generationConfig": {
                "temperature": 0.7,
                "maxOutputTokens": 100
            }
        }
        
        api_key = os.environ.get("GEMINI_API_KEY") # Use environment variable
        if not api_key:
            return jsonify({'status': 'error', 'message': 'GEMINI_API_KEY is not set in environment variables.'}), 500

        api_url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent?key={api_key}" 

        response = requests.post(api_url, headers={'Content-Type': 'application/json'}, json=payload)
        response.raise_for_status()
        
        result = response.json()
        
        if result and result.get('candidates') and result['candidates'][0].get('content') and result['candidates'][0]['content'].get('parts'):
            generated_text = result['candidates'][0]['content']['parts'][0]['text']
            return jsonify({'status': 'success', 'description': generated_text})
        else:
            return jsonify({'status': 'error', 'message': 'Failed to get a valid response from the AI. Response structure unexpected.'}), 500

    except requests.exceptions.RequestException as e:
        print(f"Error calling AI API: {e}")
        return jsonify({'status': 'error', 'message': f'Error contacting AI: {e}'}), 500
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
        return jsonify({'status': 'error', 'message': f'Something unexpected happened: {e}'}), 500

@app.route('/generate_attendance_insights', methods=['POST'])
@login_required
def generate_attendance_insights():
    api_key = os.environ.get("GEMINI_API_KEY")
    if not api_key:
        return jsonify({'status': 'error', 'message': 'AI Service is not configured.'}), 500

    filters = request.get_json()
    log_entries = get_filtered_attendance_data(filters)

    if not log_entries:
        return jsonify({'status': 'error', 'message': 'No data found for the selected filters to generate insights.'})

    present_count = sum(1 for log in log_entries if log['status'] == 'Present')
    absent_count = sum(1 for log in log_entries if log['status'] == 'Absent')
    total_students = present_count + absent_count

    if total_students == 0:
         return jsonify({'status': 'error', 'message': 'No students found for this filter combination.'})

    present_percentage = (present_count / total_students) * 100
    absent_students_list = [f"{log['student_name']} ({log['student_roll_no']})" for log in log_entries if log['status'] == 'Absent']
    
    prompt = f"""
    As an academic assistant, analyze the following daily attendance report and provide a concise summary.

    Report Details:
    - Course/Period: {filters.get('period_name', 'N/A')}
    - Date: {filters.get('date', 'N/A')}
    - Branch: {filters.get('branch', 'All')}
    - Semester: {filters.get('semester', 'All')}
    - Total Students: {total_students}
    - Present: {present_count}
    - Absent: {absent_count}
    - Absent Students: {', '.join(absent_students_list) if absent_students_list else 'None'}
    - Attendance Percentage: {present_percentage:.2f}%

    Task:
    1.  Start with a headline summarizing the attendance (e.g., "High Turnout" or "Low Attendance Alert").
    2.  In one paragraph, briefly describe the situation. Mention the attendance percentage.
    3.  Provide 1-2 actionable suggestions for the professor or administrator based on the data. For example, if attendance is low, suggest following up with absent students. If attendance is high, suggest acknowledging the class's engagement. Keep it brief.
    
    Generate the response in Markdown format."""

    api_url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-pro:generateContent?key={api_key}"
    payload = {"contents": [{"parts": [{"text": prompt}]}]}
    
    try:
        response = requests.post(api_url, json=payload)
        response.raise_for_status()
        generated_text = response.json()['candidates'][0]['content']['parts'][0]['text']
        return jsonify({'status': 'success', 'insights': generated_text.strip()})
    except Exception as e:
        return jsonify({'status': 'error', 'message': f'Error from AI service: {e}'}), 500
    except Exception as e:
        return jsonify({'status': 'error', 'message': f'Error from AI service: {e}'}), 500

@app.route('/live')
@login_required
@role_required(roles=['admin', 'professor'])
def live_attendance():
    """Renders the live attendance camera feed page."""
    print(f"DEBUG: Session content when /live HTML is requested: {session.get('attendance_config')}")
    return render_template('live.html')


@app.route('/configure', methods=['GET', 'POST'])
@login_required
@role_required(roles=['admin', 'professor'])
def configure_session():
    current_config = session.get('attendance_config')
    if request.method == 'POST':
        selected_branches = request.form.getlist('branches')
        selected_semester = request.form.get('semester')
        selected_subject = request.form.get('subject') 

        if not selected_branches or not selected_semester:
            flash("You must select at least one branch and a semester.", "warning")
            return redirect(url_for('configure_session'))

        session['attendance_config'] = {
            'branches': selected_branches,
            'semester': int(selected_semester),
            'subject': selected_subject
        }
        flash(f"Session configured for Branch: {', '.join(selected_branches)}, Semester: {selected_semester}. Starting live feed.", "success")
        return redirect(url_for('live_attendance'))

    available_branches = get_distinct_from_db('branch', 'students')
    available_semesters = get_distinct_from_db('semester', 'students')
    available_subjects = get_distinct_from_db('subject', 'students')
    
    return render_template('configure_session.html',
                           current_config=current_config,
                           available_branches=available_branches,
                           available_semesters=available_semesters,
                           available_subjects=available_subjects)


@app.route('/export_attendance', methods=['GET'])
@login_required
@role_required(roles=['admin', 'professor'])
def export_attendance():
    # Retrieve filter parameters, identical to view_attendance_log
    filter_branch = request.args.get('branch')
    filter_semester = request.args.get('semester')
    filter_period_id = request.args.get('period_id')
    filter_date_str = request.args.get('date')
    filter_status = request.args.get('status', 'all')

    current_filters = {
        'branch': filter_branch,
        'semester': filter_semester,
        'period_id': filter_period_id,
        'date': filter_date_str,
        'status': filter_status
    }

    log_entries_for_excel = get_attendance_data_for_display(current_filters)

    if not log_entries_for_excel:
        flash("No attendance records found for the selected filters to export.", "warning")
        return redirect(url_for('view_attendance_log', **current_filters)) # Pass filters back to log page

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance_Report"

    headers_map = {
        "ROLL NO": 'student_roll_no', # Changed order, removed LOG ID as absent don't have it
        "STUDENT NAME": 'student_name',
        "BRANCH": 'student_branch', # These will now be properly populated for all
        "SEMESTER": 'student_semester', # These will now be properly populated for all
        "PERIOD": 'period_name',
        "PROFESSOR": 'prof_name',
        "STATUS": 'status', # New column for Excel export
        "TIMESTAMP": 'timestamp'
    }
    
    header_titles = list(headers_map.keys())
    ws.append(header_titles)

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
    for col_idx, cell in enumerate(ws[1]):
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = header_fill
        ws.column_dimensions[cell.column_letter].width = 15 

    for row_data in log_entries_for_excel:
        row_values = []
        for header_title in header_titles:
            key = headers_map[header_title]
            value = row_data.get(key, 'N/A')
            row_values.append(value)
        ws.append(row_values)

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                cell_value_str = str(cell.value) if cell.value is not None else ""
                if len(cell_value_str) > max_length:
                    max_length = len(cell_value_str)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.1
        ws.column_dimensions[column].width = adjusted_width

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = "attendance_report"
    if current_filters['branch'] and current_filters['branch'] != 'all':
        filename += f"_Branch_{current_filters['branch']}"
    if current_filters['semester'] and current_filters['semester'] != 'all':
        filename += f"_Sem_{current_filters['semester']}"
    if current_filters['period_id'] and current_filters['period_id'] != 'all':
        # Get period name from ID for filename
        conn = get_db()
        period_name_for_filename = conn.execute("SELECT period_name FROM schedule WHERE period_id = ?", (current_filters['period_id'],)).fetchone()
        conn.close()
        if period_name_for_filename:
            filename += f"_{period_name_for_filename['period_name'].replace(' ', '_').replace(':', '')}"
        else:
            filename += f"_PeriodID_{current_filters['period_id']}"
    if current_filters['date']:
        filename += f"_Date_{current_filters['date'].replace('-', '')}"
    if current_filters['status'] and current_filters['status'] != 'all':
        filename += f"_{current_filters['status']}"
    
    filename += f"_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )


@socketio.on('image_from_client')
def handle_image_from_client(data):
    global global_blink_counter, global_blink_total, global_last_ear, last_liveness_check_time

    start_total_time = time.time()

    detected_age_for_display = 'N/A'
    detected_gender_for_display = 'N/A'
    liveness_status = 'N/A' # New: Initialize liveness status
    
    try:
        # Decode image from base64
        img_data = base64.b64decode(data.split(',')[1])
        np_arr = np.frombuffer(img_data, np.uint8)
        img = cv2.imdecode(np_arr, cv2.IMREAD_COLOR)
        
        # Convert BGR to RGB for face_recognition library
        rgb_img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        
        # Get face locations and encodings
        face_locations = face_recognition.face_locations(rgb_img)
        encodings = face_recognition.face_encodings(rgb_img, face_locations)

        conn = db_connect()
        try:
            results = [] 
            for i, (encoding, location) in enumerate(zip(encodings, face_locations)):
                name, roll_no, student_gender_from_db, student_age_from_db = None, None, 'N/A', 'N/A' 
                current_status = 'unknown' 

                # Convert dlib bounding box format (top, right, bottom, left) to (left, top, width, height)
                # for dlib.rectangle object (dlib predictor takes dlib.rectangle or a tuple)
                # face_recognition gives (top, right, bottom, left)
                (top, right, bottom, left) = location
                dlib_rect = dlib.rectangle(left, top, right, bottom) # left, top, right, bottom

                # --- Perform Liveness Check (Blink Detection) ---
                if predictor and (time.time() - last_liveness_check_time) > LIVENESS_CHECK_COOLDOWN:
                    
                    # Get grayscale image for dlib landmark prediction (it's faster)
                    gray_img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY) # Convert original BGR img to grayscale

                    # Predict facial landmarks
                    shape = predictor(gray_img, dlib_rect)
                    # Convert dlib shape to numpy array (x, y)-coordinates
                    shape = np.array([(shape.part(j).x, shape.part(j).y) for j in range(68)])

                    # Extract eye landmarks (points 36-41 for left eye, 42-47 for right eye)
                    left_eye = shape[36:42]
                    right_eye = shape[42:48]

                    # Calculate EAR for both eyes
                    left_ear = eye_aspect_ratio(left_eye)
                    right_ear = eye_aspect_ratio(right_eye)

                    # Average the eye aspect ratio together for both eyes
                    ear = (left_ear + right_ear) / 2.0

                    # Update global blink detection state
                    if ear < EYE_AR_THRESH:
                        global_blink_counter += 1
                    else:
                        # If eye is open and counter reached threshold, it's a blink
                        if global_blink_counter >= EYE_AR_CONSEC_FRAMES:
                            global_blink_total += 1
                            liveness_status = 'Blink Detected!' # Or 'Liveness: PASS'
                            last_liveness_check_time = time.time() # Reset cooldown
                        global_blink_counter = 0 # Reset counter

                    global_last_ear = ear # Store last EAR value

                    # Basic thresholding for static liveness check (if no blink happens for a while)
                    # This is just an example, a full liveness system needs more sophisticated logic.
                    if liveness_status == 'N/A': # If no blink detected yet for this cycle
                        if ear > EYE_AR_THRESH * 1.5: # Eye is wide open, less likely to be print
                             liveness_status = 'Eyes Open' # Indicate active
                        elif ear < EYE_AR_THRESH / 2: # Eye is very closed, possibly not natural
                             liveness_status = 'Eyes Closed/Low EAR'
                        else:
                             liveness_status = 'No Blink Yet' # Default if nothing conclusive
                else:
                    liveness_status = 'Liveness Cooldown' # Not checking due to cooldown or predictor not loaded


                # --- Face Comparison and Attendance Logic (unchanged) ---
                start_compare_time = time.time()
                face_distances = face_recognition.face_distance(known_face_encodings, encoding)
                best_match_index = np.argmin(face_distances)
                best_match_distance = face_distances[best_match_index]
                end_compare_time = time.time()
                
                CONFIDENCE_THRESHOLD = 0.5 
                
                if best_match_distance < CONFIDENCE_THRESHOLD:
                    roll_no = known_face_roll_numbers[best_match_index]

                    student = conn.execute("SELECT name, branch, semester, gender, age FROM students WHERE roll_no = ?", (roll_no,)).fetchone()
                    
                    if student:
                        name = student['name']
                        student_gender_from_db = student['gender'] if student['gender'] else 'N/A'
                        student_age_from_db = student['age'] if student['age'] else 'N/A'
                        
                        detected_gender_for_display = student_gender_from_db
                        detected_age_for_display = student_age_from_db

                        current_time_sec = time.time()
                        session_config = session.get('attendance_config')
                        
                        if session_config:
                            config_branches = session_config.get('branches', [])
                            config_semester = session_config.get('semester')
                            config_subject = session_config.get('subject')

                            student_meets_criteria = (
                                (not config_branches or student['branch'] in config_branches) and
                                (config_semester is None or student['semester'] == config_semester) and
                                (not config_subject or student.get('subject') == config_subject)
                            )
                            
                            if student_meets_criteria:
                                current_period = None
                                current_prof_name = None
                                local_now_time = datetime.now(LOCAL_TIMEZONE).time()
                                periods = conn.execute("SELECT period_name, start_time, end_time, prof_name FROM schedule").fetchall()
                                for period_rec in periods:
                                    period_start = datetime.strptime(period_rec['start_time'], '%H:%M').time()
                                    period_end = datetime.strptime(period_rec['end_time'], '%H:%M').time()
                                    if period_start <= local_now_time <= period_end:
                                        current_period = period_rec['period_name']
                                        current_prof_name = period_rec['prof_name']
                                        break
                                
                                if current_period:
                                    local_now_datetime = datetime.now(LOCAL_TIMEZONE)
                                    timestamp_str = local_now_datetime.strftime("%Y-%m-%d %H:%M:%S")
                                    today_str = local_now_datetime.strftime("%Y-%m-%d")

                                    existing_log = conn.execute("SELECT 1 FROM attendance_logs WHERE student_roll_no = ? AND period_name = ? AND date(timestamp) = ?", (roll_no, current_period, today_str)).fetchone()
                                    
                                    if not existing_log:
                                        if roll_no not in last_logged_time or (current_time_sec - last_logged_time[roll_no]) > COOLDOWN_PERIOD:
                                            # --- Crucial: Only log if Liveness is successful! ---
                                            # For simplicity, let's tie it to a blink detection
                                            # A real system needs more robust liveness_status == 'PASS'
                                            if liveness_status == 'Blink Detected!': # Only log if blink detected
                                                conn.execute("INSERT INTO attendance_logs (student_roll_no, student_name, period_name, prof_name, timestamp) VALUES (?, ?, ?, ?, ?)",
                                                            (roll_no, name, current_period, current_prof_name, timestamp_str))
                                                conn.commit()
                                                last_logged_time[roll_no] = current_time_sec
                                                current_status = 'logged'
                                                socketio.emit('new_log_entry', {
                                                    'name': name,
                                                    'roll_no': roll_no,
                                                    'gender': detected_gender_for_display, 
                                                    'age': detected_age_for_display,       
                                                    'time': local_now_datetime.strftime('%H:%M:%S')
                                                })
                                            else: # Liveness failed, or not active
                                                current_status = 'Liveness Check Failed' # Or 'Awaiting Liveness'
                                        else:
                                            current_status = 'cooldown'
                                    else:
                                        current_status = 'already_logged'
                                else:
                                    current_status = 'no_active_period'
                            else:
                                current_status = 'filter_mismatch'
                        else:
                            current_status = 'session_not_configured'
                    else:
                        name = "Unknown"
                        roll_no = None
                        current_status = 'db_mismatch' 

                # Append results for this face to the list for the client
                results.append({
                    'name': name,
                    'roll_no': roll_no,
                    'location': location,
                    'status': current_status,
                    'gender': detected_gender_for_display, 
                    'age': detected_age_for_display,       
                    'distance': best_match_distance,
                    'liveness_status': liveness_status # NEW: Pass liveness status to frontend
                })
            
            socketio.emit('recognition_result', {
                'status': 'success', 
                'results': results,
                'period': 'N/A' 
            })

        except Exception as e:
            print(f"Error in handle_image_from_client (inner try block): {e}")
            socketio.emit('recognition_result', {'status': 'error', 'message': str(e), 'results': []})
        finally:
            conn.close() 
        
        end_total_time = time.time()
        print(f"Total frame processing time: {(end_total_time - start_total_time)*1000:.2f} ms")

    except Exception as e:
        print(f"Top-level error in handle_image_from_client: {e}")
        socketio.emit('recognition_result', {'status': 'error', 'message': f"Top-level processing error: {e}", 'results': []})
# --- Main Execution ---
if __name__ == '__main__':
    print("Starting Flask-SocketIO server on http://127.0.0.1:5000")
    socketio.run(app, debug=True, port=5000)